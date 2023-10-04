from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import time
import os

class RanDown():
    def __init__(self):
        self.filename = None

    def open_file(self,filename):
        self.filename = filename
        wb = load_workbook(self.filename)
        # sh = wb.get_sheet_by_name("Sheet1") #用这个报错
        sh = wb["Sheet1"]
        rows = sh.max_row
        columns = sh.max_column
        print(rows)
        print(columns)
        self.url = []
        for index in range(1,rows+1):
            print("index=",index)
            url1 = sh.cell(row=index,column=2).value
            self.url.append(url1)
        print(self.url)
        return self.url

    def auto_get(self):
        # self.driver = webdriver.Firefox()
        '''
        设置好系统的环境变量，以后每次调用chromedriver.exe时都不需要多余的代码来设置系统环境变量路径；
        在代码中设置环境变量（亲测这个还是比较好用的）
         '''
        abspath = os.path.abspath(r"C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe")
        self.driver = webdriver.Chrome(abspath)
        self.driver.maximize_window()
        for sub_url in self.url:
            print(sub_url)
            self.driver.get(sub_url)
            self.driver.implicitly_wait(20)
            self.driver.find_elements_by_class_name("btn")[1].click()  #用这个和xpath一样
            # self.driver.find_element_by_xpath("/html/body/form/table/tbody/tr[2]/td/li/ul/input[2]").click()
        print("Download finished.")


if __name__ == '__main__':
    randown = RanDown()
    randown.open_file("OpenRanDown.xlsx")
    randown.auto_get()
